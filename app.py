from __future__ import annotations

import os
import base64
import hashlib
import secrets
import re
import shutil
import sqlite3
import subprocess
from zipfile import ZIP_DEFLATED, ZipFile
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

import holidays
from cryptography.fernet import Fernet, InvalidToken
from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from docx import Document
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

DB_PATH = Path(os.environ.get("URLAUBSPLANER_DB", "instance/urlaubsplaner.sqlite3"))
VBA_PROJECT_BASE64_PATH = Path(__file__).with_name("assets") / "vbaProject.bin.b64"
ROLE_ORDER = ["admin", "ausbilder", "putzchef", "azubi", "normal", "desksharing"]
ROLE_RANK = {role: index for index, role in enumerate(ROLE_ORDER)}
IGNORED_IMPORT_NAMES = {
    "monat",
    "kw",
    "ferienzeit",
    "geplant oder beantragt",
    "genehmigt",
    "feiertag",
    "grundkurs",
    "berufsschule",
    "ausbildungsmesse",
    "pv+ba+ap",
    "ausbildung",
    "weihnachtsputz",
    "kurzarbeit",
    "weiterbildung",
    "kein arbeitstag",
}
ENTRY_TYPES = {
    "UB": ("Urlaub geplant/beantragt", "FFFF00"),
    "UG": ("Genehmigter Urlaub", "90EE90"),
    "PV": ("PV+BA+AP", "B084CC"),
    "BA": ("PV+BA+AP", "B084CC"),
    "AP": ("PV+BA+AP", "B084CC"),
    "AZ": ("Ausbildung", "006100"),
    "KA": ("Kurzarbeit", "FF0000"),
    "WB": ("Weiterbildung", "FFA500"),
    "AM": ("Ausbildungsmesse", "FF69B4"),
    "BS": ("Berufsschule", "40E0D0"),
    "GK": ("Grundkurs", "C4A484"),
    "WP": ("Weihnachtsputz", "654321"),
    "KAT": ("Kein Arbeitstag", "000000"),
    "KR": ("Krank", "90EE90"),
}
IMPORT_VALUE_TO_CODE = {
    "UrlbGplntOdrBntrgt": "UB",
    "UrlbGnhmgt": "UG",
    "Frtg": None,
    "Grndkrs": "GK",
    "Brfsschl": "BS",
    "Asbldngsmss": "AM",
    "PvBaAp": "PV",
    "Asbldng": "AZ",
    "Whnchtsptz": "WP",
    "Krzrbt": "KA",
    "Wtrbldng": "WB",
    "KnArbtstg": "KAT",
}
IMPORT_VALUE_TO_CODE_NORMALIZED = {
    key.casefold(): value for key, value in IMPORT_VALUE_TO_CODE.items()
}
IMPORT_LEGEND_LABEL_TO_CODE = {
    "geplant oder beantragt": "UB",
    "genehmigt": "UG",
    "feiertag": None,
    "grundkurs": "GK",
    "berufsschule": "BS",
    "ausbildungsmesse": "AM",
    "pv+ba+ap": "PV",
    "ausbildung": "AZ",
    "weihnachtsputz": "WP",
    "kurzarbeit": "KA",
    "weiterbildung": "WB",
    "kein arbeitstag": "KAT",
}
SPECIAL_KR = "00FE43"
HOLIDAY_COLOR = "ADD8E6"
VACATION_COLOR = "BDD7EE"
ALLOWED_BY_ROLE = {
    "normal": {"UB", "UG"},
    "desksharing": {"UB", "UG"},
    "azubi": {"UB", "UG", "BS"},
    "putzchef": {"UB", "UG", "WP"},
    "ausbilder": {"UB", "UG", "AM", "BS", "GK", "WP"},
    "admin": set(ENTRY_TYPES),
}
GERMAN_WEEKDAYS = ["Mo", "Di", "Mi", "Do", "Fr"]
DESKSHARING_STATUSES = {
    "Anwesend": "72C472",
    "Homeoffice": "74A9E6",
    "Abwesend": "D9D9D9",
}
DESKSHARING_WEEKDAYS = {
    "mo": 1,
    "montag": 1,
    "di": 2,
    "dienstag": 2,
    "mi": 3,
    "mittwoch": 3,
    "do": 4,
    "donnerstag": 4,
    "fr": 5,
    "freitag": 5,
}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-change-me")
app.config["UPLOAD_FOLDER"] = "instance/uploads"
app.config["PROFILE_UPLOAD_FOLDER"] = "instance/profile-images"
login_manager = LoginManager(app)
login_manager.login_view = "login"


class User(UserMixin):
    def __init__(self, row: sqlite3.Row, roles: list[str]):
        self.id = str(row["id"])
        self.first_name = row["first_name"]
        self.last_name = row["last_name"]
        self.username = row["username"]
        self.password_hash = row["password_hash"]
        self.must_change_password = bool(row["must_change_password"])
        self.profile_image = row["profile_image"]
        self.roles = roles

    def has_role(self, role: str) -> bool:
        return role in self.roles

    @property
    def display_name(self) -> str:
        return f"{self.first_name} {self.last_name}"


def db() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                first_name TEXT NOT NULL,
                last_name TEXT NOT NULL,
                password_hash TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS user_roles (
                user_id INTEGER NOT NULL,
                role TEXT NOT NULL,
                PRIMARY KEY (user_id, role),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                entry_date TEXT NOT NULL,
                code TEXT NOT NULL,
                created_by INTEGER,
                UNIQUE(user_id, entry_date, code),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS desksharing_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                entry_date TEXT NOT NULL,
                status TEXT NOT NULL,
                created_by INTEGER,
                UNIQUE(user_id, entry_date),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS user_matrix_order (
                matrix_name TEXT NOT NULL,
                user_id INTEGER NOT NULL,
                position INTEGER NOT NULL,
                PRIMARY KEY (matrix_name, user_id),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            """
        )
        columns = {row["name"] for row in conn.execute("PRAGMA table_info(users)")}
        migrations = {
            "must_change_password": "INTEGER NOT NULL DEFAULT 0",
            "initial_password_encrypted": "TEXT",
            "initial_data_acknowledged": "INTEGER NOT NULL DEFAULT 1",
            "email": "TEXT",
            "birth_date": "TEXT",
            "profile_image": "TEXT",
        }
        for column, definition in migrations.items():
            if column not in columns:
                conn.execute(f"ALTER TABLE users ADD COLUMN {column} {definition}")
        if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
            user_id = create_user(
                conn, "admin", "Admin", "Benutzer", "admin", ["admin"]
            )
            conn.execute(
                "INSERT OR IGNORE INTO entries(user_id, entry_date, code, created_by) VALUES (?, ?, ?, ?)",
                (user_id, date.today().isoformat(), "UG", user_id),
            )


def create_user(
    conn, username: str, first: str, last: str, password: str, roles: list[str]
) -> int:
    cur = conn.execute(
        "INSERT INTO users(username, first_name, last_name, password_hash) VALUES (?, ?, ?, ?)",
        (username, first, last, generate_password_hash(password)),
    )
    user_id = cur.lastrowid
    for role in normalized_roles(roles):
        conn.execute(
            "INSERT INTO user_roles(user_id, role) VALUES (?, ?)", (user_id, role)
        )
    return user_id


def initial_password() -> str:
    """Create an eight-character password without ambiguous characters."""
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(8))


def credential_cipher() -> Fernet:
    key = hashlib.sha256(app.config["SECRET_KEY"].encode()).digest()
    return Fernet(base64.urlsafe_b64encode(key))


def create_initial_user(
    conn, username: str, first: str, last: str, roles: list[str]
) -> tuple[int, str]:
    password = initial_password()
    encrypted = credential_cipher().encrypt(password.encode()).decode()
    cur = conn.execute(
        """INSERT INTO users(
            username, first_name, last_name, password_hash,
            must_change_password, initial_password_encrypted,
            initial_data_acknowledged
        ) VALUES (?, ?, ?, ?, 1, ?, 0)""",
        (username, first, last, generate_password_hash(password), encrypted),
    )
    user_id = cur.lastrowid
    for role in normalized_roles(roles):
        conn.execute(
            "INSERT INTO user_roles(user_id, role) VALUES (?, ?)", (user_id, role)
        )
    return user_id, password


def reveal_initial_password(row: sqlite3.Row) -> str:
    return (
        credential_cipher().decrypt(row["initial_password_encrypted"].encode()).decode()
    )


def initial_link(username: str, password: str) -> str:
    payload = f"{username}\0{password}".encode()
    token = credential_cipher().encrypt(payload).decode()
    return f"https://urlaub.extrahelden.de/initial-login?token={token}"


def normalized_roles(roles: list[str]) -> list[str]:
    valid = [role for role in ROLE_ORDER if role in roles]
    return valid or ["normal"]


@login_manager.user_loader
def load_user(user_id: str) -> User | None:
    with db() as conn:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not row:
            return None
        roles = [
            r["role"]
            for r in conn.execute(
                "SELECT role FROM user_roles WHERE user_id = ?", (user_id,)
            )
        ]
        return User(row, roles)


def weekdays_between(start: date, end: date) -> list[date]:
    days = []
    current = start
    while current <= end:
        if current.weekday() < 5:
            days.append(current)
        current += timedelta(days=1)
    return days


def band_range(user: User, year: int | None = None) -> tuple[date, date]:
    today = date.today()
    base_year = year or today.year
    if user.has_role("admin") or user.has_role("ausbilder"):
        return date(base_year - 1, 7, 2), date(base_year + 1, 7, 2)
    return date(base_year, 1, 1), date(base_year, 12, 31)


def current_week_monday() -> date:
    today = date.today()
    return today - timedelta(days=today.weekday())


def get_users(conn) -> list[sqlite3.Row]:
    rows = conn.execute(
        """
        SELECT u.*, COALESCE(MIN(CASE ur.role
            WHEN 'admin' THEN 0 WHEN 'ausbilder' THEN 1 WHEN 'putzchef' THEN 2
            WHEN 'azubi' THEN 3 ELSE 4 END), 4) AS rank
        FROM users u LEFT JOIN user_roles ur ON u.id = ur.user_id
        GROUP BY u.id ORDER BY rank, u.last_name COLLATE NOCASE, u.first_name COLLATE NOCASE
        """
    ).fetchall()
    return rows


def roles_for_users(conn) -> dict[int, list[str]]:
    result = defaultdict(list)
    for row in conn.execute("SELECT user_id, role FROM user_roles ORDER BY user_id"):
        result[row["user_id"]].append(row["role"])
    return result


def get_desksharing_users(conn) -> list[sqlite3.Row]:
    desksharing_ids = {
        row["user_id"]
        for row in conn.execute(
            "SELECT user_id FROM user_roles WHERE role = 'desksharing'"
        )
    }
    return [user for user in get_users(conn) if user["id"] in desksharing_ids]


def order_matrix_users(
    conn, users: list[sqlite3.Row], matrix_name: str
) -> list[sqlite3.Row]:
    positions = {
        row["user_id"]: row["position"]
        for row in conn.execute(
            "SELECT user_id, position FROM user_matrix_order WHERE matrix_name = ?",
            (matrix_name,),
        )
    }
    fallback = len(positions) + len(users)
    original_positions = {user["id"]: index for index, user in enumerate(users)}
    return sorted(
        users,
        key=lambda user: (
            positions.get(user["id"], fallback + original_positions[user["id"]]),
            original_positions[user["id"]],
        ),
    )


def visible_code(
    viewer: User, target_user_id: int, target_roles: list[str], code: str
) -> tuple[str, str]:
    color = ENTRY_TYPES[code][1]
    if code == "KR":
        if (
            viewer.has_role("admin")
            or target_user_id == int(viewer.id)
            or (viewer.has_role("ausbilder") and "azubi" in target_roles)
        ):
            color = SPECIAL_KR
    return code if viewer.has_role("admin") else "", color


def month_spans(days: list[date]) -> list[dict[str, int | str]]:
    spans = []
    previous = None
    for day in days:
        key = (day.year, day.month)
        if key != previous:
            spans.append({"label": day.strftime("%b"), "span": 1})
            previous = key
        else:
            spans[-1]["span"] += 1
    return spans


def year_spans(days: list[date]) -> list[dict[str, int | str]]:
    spans = []
    previous = None
    for day in days:
        if day.year != previous:
            spans.append({"label": str(day.year), "span": 1})
            previous = day.year
        else:
            spans[-1]["span"] += 1
    return spans


def cell_style_and_tooltip(
    viewer: User,
    target_user_id: int,
    target_roles: list[str],
    day: date,
    codes: list[str],
) -> tuple[str, str]:
    if not codes:
        return "", ""
    code = codes[-1]
    label, default_color = ENTRY_TYPES[code]
    _, color = visible_code(viewer, target_user_id, target_roles, code)
    if viewer.has_role("admin"):
        entry_text = f"{code} – {label}"
    elif code == "KR" and color != SPECIAL_KR:
        entry_text = "Abwesend"
    else:
        entry_text = label
    return default_color if code != "KR" else color, f"{day:%d.%m.%Y}: {entry_text}"


def school_vacations_bavaria(years: set[int]) -> dict[str, str]:
    # Ferien nach bayerischem Ferienkalender; feste Berechnung für Schuljahre 2025-2027.
    ranges = {
        2025: [
            ("2025-03-03", "2025-03-07"),
            ("2025-04-14", "2025-04-25"),
            ("2025-06-10", "2025-06-20"),
            ("2025-08-01", "2025-09-15"),
            ("2025-11-03", "2025-11-07"),
            ("2025-12-22", "2026-01-05"),
        ],
        2026: [
            ("2025-12-22", "2026-01-05"),
            ("2026-02-16", "2026-02-20"),
            ("2026-03-30", "2026-04-10"),
            ("2026-05-26", "2026-06-05"),
            ("2026-08-03", "2026-09-14"),
            ("2026-11-02", "2026-11-06"),
            ("2026-12-24", "2027-01-08"),
        ],
        2027: [
            ("2026-12-24", "2027-01-08"),
            ("2027-02-08", "2027-02-12"),
            ("2027-03-22", "2027-04-02"),
            ("2027-05-18", "2027-05-28"),
            ("2027-08-02", "2027-09-13"),
            ("2027-11-02", "2027-11-05"),
            ("2027-12-24", "2028-01-07"),
        ],
    }
    result = {}
    for year in years:
        for start_s, end_s in ranges.get(year, []):
            for day in weekdays_between(
                date.fromisoformat(start_s), date.fromisoformat(end_s)
            ):
                result[day.isoformat()] = "Ferien"
    return result


def bavarian_holidays(years: set[int]) -> dict[str, str]:
    by = holidays.Germany(subdiv="BY", years=years, language="de")
    return {d.isoformat(): name for d, name in by.items() if d.weekday() < 5}


def allowed_codes(user: User) -> list[str]:
    codes = set()
    for role in user.roles:
        codes |= ALLOWED_BY_ROLE.get(role, set())
    return [code for code in ENTRY_TYPES if code in codes]


def may_edit(user: User, target_id: int, code: str) -> bool:
    return user.has_role("admin") or (
        target_id == int(user.id) and code in allowed_codes(user)
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        with db() as conn:
            row = conn.execute(
                "SELECT * FROM users WHERE username = ?", (request.form["username"],)
            ).fetchone()
        if row and check_password_hash(row["password_hash"], request.form["password"]):
            login_user(load_user(str(row["id"])))
            if row["must_change_password"]:
                return redirect(url_for("initial_login"))
            return redirect(url_for("index"))
        flash("Ungültige Zugangsdaten.")
    return render_template("login.html")


@app.route("/initial-login", methods=["GET", "POST"])
def initial_login():
    username = ""
    password = ""
    token = request.args.get("token", "")
    if token:
        try:
            username, password = (
                credential_cipher().decrypt(token.encode()).decode().split("\0", 1)
            )
        except (InvalidToken, ValueError):
            flash("Der Initial-Link ist ungültig oder abgelaufen.")
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["initial_password"]
        new_password = request.form["new_password"]
        if new_password != request.form["password_confirmation"]:
            flash("Die neuen Passwörter stimmen nicht überein.")
        elif len(new_password) < 8:
            flash("Das neue Passwort muss mindestens 8 Zeichen lang sein.")
        else:
            with db() as conn:
                row = conn.execute(
                    "SELECT * FROM users WHERE username = ?", (username,)
                ).fetchone()
                if (
                    row
                    and row["must_change_password"]
                    and check_password_hash(row["password_hash"], password)
                ):
                    conn.execute(
                        """UPDATE users SET password_hash = ?, must_change_password = 0,
                           initial_password_encrypted = NULL WHERE id = ?""",
                        (generate_password_hash(new_password), row["id"]),
                    )
                    logout_user()
                    flash("Passwort gespeichert. Du kannst dich jetzt anmelden.")
                    return redirect(url_for("login"))
            flash("Benutzername oder Initialpasswort ist ungültig.")
    return render_template(
        "initial-login.html", username=username, initial_password=password
    )


@app.route("/logout")
@login_required
def logout():
    logout_user()
    session.pop("pending_initial_user_id", None)
    return redirect(url_for("login"))


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    with db() as conn:
        row = conn.execute(
            "SELECT * FROM users WHERE id = ?", (current_user.id,)
        ).fetchone()
        if request.method == "POST":
            if request.form.get("action") == "delete-profile-image":
                if row["profile_image"]:
                    image_path = (
                        Path(app.config["PROFILE_UPLOAD_FOLDER"]) / row["profile_image"]
                    )
                    image_path.unlink(missing_ok=True)
                    conn.execute(
                        "UPDATE users SET profile_image = NULL WHERE id = ?",
                        (current_user.id,),
                    )
                flash("Profilbild gelöscht.")
                return redirect(url_for("profile"))
            email = (row["email"] or request.form.get("email", "")).strip()
            if not email:
                flash("Die E-Mail-Adresse ist eine Pflichtangabe.")
                return render_template("profile.html", profile=row)
            birth_date = request.form.get("birth_date") or None
            if birth_date:
                try:
                    date.fromisoformat(birth_date)
                except ValueError:
                    flash("Das Geburtsdatum ist ungültig.")
                    return render_template("profile.html", profile=row)
            image_name = row["profile_image"]
            image = request.files.get("profile_image")
            if image and image.filename:
                suffix = Path(secure_filename(image.filename)).suffix.lower()
                if suffix not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
                    flash("Bitte lade ein Bild im Format JPG, PNG, WEBP oder GIF hoch.")
                    return render_template("profile.html", profile=row)
                Path(app.config["PROFILE_UPLOAD_FOLDER"]).mkdir(
                    parents=True, exist_ok=True
                )
                image_name = f"user-{current_user.id}{suffix}"
                image.save(Path(app.config["PROFILE_UPLOAD_FOLDER"]) / image_name)
            conn.execute(
                "UPDATE users SET email = ?, birth_date = ?, profile_image = ? WHERE id = ?",
                (email, birth_date, image_name, current_user.id),
            )
            flash("Profil gespeichert.")
            return redirect(url_for("profile"))
    return render_template("profile.html", profile=row)


@app.route("/profile-image/<path:filename>")
@login_required
def profile_image(filename: str):
    return send_from_directory(app.config["PROFILE_UPLOAD_FOLDER"], filename)


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        with db() as conn:
            row = conn.execute(
                "SELECT * FROM users WHERE id = ?", (current_user.id,)
            ).fetchone()
            new_password = request.form["new_password"]
            if not check_password_hash(
                row["password_hash"], request.form["current_password"]
            ):
                flash("Das aktuelle Passwort ist falsch.")
            elif new_password != request.form["password_confirmation"]:
                flash("Die neuen Passwörter stimmen nicht überein.")
            elif len(new_password) < 8:
                flash("Das neue Passwort muss mindestens 8 Zeichen lang sein.")
            else:
                conn.execute(
                    "UPDATE users SET password_hash = ? WHERE id = ?",
                    (generate_password_hash(new_password), current_user.id),
                )
                flash("Passwort geändert.")
                return redirect(url_for("profile"))
    return render_template("change-password.html")


@app.route("/")
@login_required
def index():
    year = int(request.args.get("year", date.today().year))
    start, end = band_range(current_user, year)
    days = weekdays_between(start, end)
    today_monday = current_week_monday().isoformat()
    with db() as conn:
        users = order_matrix_users(conn, get_users(conn), "vacation")
        role_map = roles_for_users(conn)
        entries = conn.execute(
            "SELECT * FROM entries WHERE entry_date BETWEEN ? AND ?",
            (start.isoformat(), end.isoformat()),
        ).fetchall()
    entry_map = defaultdict(list)
    for e in entries:
        entry_map[(e["user_id"], e["entry_date"])].append(e["code"])
    years = {d.year for d in days}
    return render_template(
        "index.html",
        users=users,
        role_map=role_map,
        days=days,
        entries=entry_map,
        entry_types=ENTRY_TYPES,
        holidays=bavarian_holidays(years),
        vacations=school_vacations_bavaria(years),
        holiday_color=HOLIDAY_COLOR,
        vacation_color=VACATION_COLOR,
        allowed_codes=allowed_codes(current_user),
        today_monday=today_monday,
        visible_code=visible_code,
        cell_style_and_tooltip=cell_style_and_tooltip,
        month_spans=month_spans(days),
        year_spans=year_spans(days),
        year=year,
        german_weekdays=GERMAN_WEEKDAYS,
    )


@app.post("/entry")
@login_required
def entry():
    target_id = int(request.form["user_id"])
    entry_date = request.form["date"]
    code = request.form["code"]
    if code not in ENTRY_TYPES or not may_edit(current_user, target_id, code):
        abort(403)
    with db() as conn:
        if request.form.get("action") == "delete":
            conn.execute(
                "DELETE FROM entries WHERE user_id = ? AND entry_date = ? AND code = ?",
                (target_id, entry_date, code),
            )
        else:
            conn.execute(
                "INSERT OR IGNORE INTO entries(user_id, entry_date, code, created_by) VALUES (?, ?, ?, ?)",
                (target_id, entry_date, code, current_user.id),
            )
    return redirect(request.referrer or url_for("index"))


@app.post("/bulk-entry")
@login_required
def bulk_entry():
    payload = request.get_json(silent=True) or {}
    cells = payload.get("cells", [])
    code = payload.get("code")
    delete = payload.get("delete", False)
    if not cells:
        return jsonify({"updated": 0})
    if not delete and code not in ENTRY_TYPES:
        abort(400)
    updated = 0
    with db() as conn:
        for cell in cells:
            target_id = int(cell["user_id"])
            entry_date = cell["date"]
            if delete:
                removable_codes = (
                    ENTRY_TYPES
                    if current_user.has_role("admin")
                    else allowed_codes(current_user)
                )
                for removable_code in removable_codes:
                    if may_edit(current_user, target_id, removable_code):
                        cur = conn.execute(
                            "DELETE FROM entries WHERE user_id = ? AND entry_date = ? AND code = ?",
                            (target_id, entry_date, removable_code),
                        )
                        updated += cur.rowcount
            elif may_edit(current_user, target_id, code):
                conn.execute(
                    "DELETE FROM entries WHERE user_id = ? AND entry_date = ?",
                    (target_id, entry_date),
                )
                conn.execute(
                    "INSERT OR IGNORE INTO entries(user_id, entry_date, code, created_by) VALUES (?, ?, ?, ?)",
                    (target_id, entry_date, code, current_user.id),
                )
                updated += 1
    return jsonify({"updated": updated})


@app.route("/members", methods=["GET", "POST"])
@login_required
def members():
    if not (current_user.has_role("admin") or current_user.has_role("ausbilder")):
        abort(403)
    credentials = None
    with db() as conn:
        if request.method == "POST":
            action = request.form["action"]
            if action == "create":
                if not current_user.has_role("admin"):
                    abort(403)
                user_id, password = create_initial_user(
                    conn,
                    request.form["username"],
                    request.form["first_name"],
                    request.form["last_name"],
                    request.form.getlist("roles"),
                )
                credentials = credential_details(conn, user_id, password)
                session["pending_initial_user_id"] = user_id
            elif action == "delete":
                if not current_user.has_role("admin"):
                    abort(403)
                conn.execute(
                    "DELETE FROM users WHERE id = ?", (request.form["user_id"],)
                )
            elif action == "roles-bulk":
                if not current_user.has_role("admin"):
                    abort(403)
                for user in get_users(conn):
                    conn.execute(
                        "DELETE FROM user_roles WHERE user_id = ?", (user["id"],)
                    )
                    roles = request.form.getlist(f"roles_{user['id']}")
                    for role in normalized_roles(roles):
                        conn.execute(
                            "INSERT INTO user_roles(user_id, role) VALUES (?, ?)",
                            (user["id"], role),
                        )
                flash("Rollen wurden aktualisiert.")
            elif action in {"reset-password", "show-initial"}:
                target = member_password_target(conn, int(request.form["user_id"]))
                if not target:
                    abort(403)
                if action == "reset-password":
                    password = reset_initial_password(conn, target["id"])
                elif (
                    target["initial_data_acknowledged"]
                    or not target["initial_password_encrypted"]
                ):
                    abort(404)
                else:
                    password = reveal_initial_password(target)
                credentials = credential_details(conn, target["id"], password)
                session["pending_initial_user_id"] = target["id"]
            if credentials is None:
                return redirect(url_for("members"))
        users = get_users(conn)
        role_map = roles_for_users(conn)
        pending_id = session.get("pending_initial_user_id")
        if credentials is None and pending_id:
            pending = member_password_target(conn, int(pending_id))
            if pending and pending["initial_password_encrypted"]:
                credentials = credential_details(
                    conn, pending["id"], reveal_initial_password(pending)
                )
    return render_template(
        "members.html",
        users=users,
        role_map=role_map,
        roles=ROLE_ORDER,
        credentials=credentials,
    )


def member_password_target(conn, user_id: int) -> sqlite3.Row | None:
    target = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target:
        return None
    if current_user.has_role("admin"):
        return target
    roles = roles_for_users(conn).get(user_id, [])
    return target if current_user.has_role("ausbilder") and "azubi" in roles else None


def reset_initial_password(conn, user_id: int) -> str:
    password = initial_password()
    encrypted = credential_cipher().encrypt(password.encode()).decode()
    conn.execute(
        """UPDATE users SET password_hash = ?, must_change_password = 1,
           initial_password_encrypted = ?, initial_data_acknowledged = 0 WHERE id = ?""",
        (generate_password_hash(password), encrypted, user_id),
    )
    return password


def credential_details(conn, user_id: int, password: str) -> dict[str, str | int]:
    row = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
    return {
        "user_id": user_id,
        "username": row["username"],
        "password": password,
        "link": initial_link(row["username"], password),
    }


@app.post("/members/acknowledge-initial-data")
@login_required
def acknowledge_initial_data():
    if not (current_user.has_role("admin") or current_user.has_role("ausbilder")):
        abort(403)
    user_id = int((request.get_json(silent=True) or {}).get("user_id", 0))
    with db() as conn:
        if not member_password_target(conn, user_id):
            abort(403)
        conn.execute(
            "UPDATE users SET initial_data_acknowledged = 1 WHERE id = ?", (user_id,)
        )
    if session.get("pending_initial_user_id") == user_id:
        session.pop("pending_initial_user_id")
    return jsonify({"ok": True})


@app.route("/years")
@login_required
def years():
    with db() as conn:
        rows = conn.execute(
            "SELECT DISTINCT substr(entry_date,1,4) AS year FROM entries WHERE user_id = ? ORDER BY year",
            (current_user.id,),
        ).fetchall()
    years_list = [r["year"] for r in rows] or [str(date.today().year)]
    return render_template("years.html", years=years_list)


@app.route("/desksharing")
@login_required
def desksharing():
    year = int(request.args.get("year", date.today().year))
    days = weekdays_between(date(year, 1, 1), date(year, 12, 31))
    with db() as conn:
        users = order_matrix_users(conn, get_desksharing_users(conn), "desksharing")
        rows = conn.execute(
            "SELECT * FROM desksharing_entries WHERE entry_date BETWEEN ? AND ?",
            (days[0].isoformat(), days[-1].isoformat()),
        ).fetchall()
    entries = {(row["user_id"], row["entry_date"]): row["status"] for row in rows}
    return render_template(
        "desksharing.html",
        year=year,
        days=days,
        users=users,
        entries=entries,
        statuses=DESKSHARING_STATUSES,
        month_spans=month_spans(days),
        year_spans=year_spans(days),
        today_monday=current_week_monday().isoformat(),
    )


@app.post("/matrix-order/<matrix_name>")
@login_required
def save_matrix_order(matrix_name: str):
    if not current_user.has_role("admin"):
        abort(403)
    if matrix_name not in {"vacation", "desksharing"}:
        abort(404)
    payload = request.get_json(silent=True) or {}
    try:
        user_ids = [int(user_id) for user_id in payload.get("user_ids", [])]
    except (TypeError, ValueError):
        abort(400)
    if len(user_ids) != len(set(user_ids)):
        abort(400)
    with db() as conn:
        visible_users = (
            get_users(conn)
            if matrix_name == "vacation"
            else get_desksharing_users(conn)
        )
        if set(user_ids) != {user["id"] for user in visible_users}:
            abort(400)
        conn.execute(
            "DELETE FROM user_matrix_order WHERE matrix_name = ?", (matrix_name,)
        )
        conn.executemany(
            "INSERT INTO user_matrix_order(matrix_name, user_id, position) VALUES (?, ?, ?)",
            [
                (matrix_name, user_id, position)
                for position, user_id in enumerate(user_ids)
            ],
        )
    return jsonify({"saved": len(user_ids)})


@app.post("/desksharing/bulk-entry")
@login_required
def desksharing_bulk_entry():
    if not current_user.has_role("admin"):
        abort(403)
    payload = request.get_json(silent=True) or {}
    cells = payload.get("cells", [])
    status = payload.get("status")
    delete = payload.get("delete", False)
    if not delete and status not in DESKSHARING_STATUSES:
        abort(400)
    with db() as conn:
        for cell in cells:
            user_id = int(cell["user_id"])
            entry_date = date.fromisoformat(cell["date"]).isoformat()
            if delete:
                conn.execute(
                    "DELETE FROM desksharing_entries WHERE user_id = ? AND entry_date = ?",
                    (user_id, entry_date),
                )
            else:
                conn.execute(
                    """INSERT INTO desksharing_entries(
                        user_id, entry_date, status, created_by
                    ) VALUES (?, ?, ?, ?)
                    ON CONFLICT(user_id, entry_date) DO UPDATE SET
                        status = excluded.status, created_by = excluded.created_by""",
                    (user_id, entry_date, status, current_user.id),
                )
    return jsonify({"updated": len(cells)})


def fill_for(viewer: User, target_id: int, roles: list[str], code: str) -> str:
    return visible_code(viewer, target_id, roles, code)[1]


@app.post("/download")
@login_required
def download():
    start_year, end_year = sorted(
        [int(request.form["start_year"]), int(request.form["end_year"])]
    )
    days = weekdays_between(date(start_year, 1, 1), date(end_year, 12, 31))
    with db() as conn:
        users = get_users(conn)
        role_map = roles_for_users(conn)
        entries = conn.execute(
            "SELECT * FROM entries WHERE entry_date BETWEEN ? AND ?",
            (days[0].isoformat(), days[-1].isoformat()),
        ).fetchall()
    entry_map = defaultdict(list)
    active_user_ids = {e["user_id"] for e in entries}
    for e in entries:
        entry_map[(e["user_id"], e["entry_date"])].append(e["code"])
    wb = Workbook()
    ws = wb.active
    ws.title = "Urlaubsübersicht"
    ws.cell(1, 1, "Name")
    for col, day in enumerate(days, 2):
        ws.cell(1, col, f"{GERMAN_WEEKDAYS[day.weekday()]} {day:%d.%m.%Y}")
    years = {d.year for d in days}
    vacations = school_vacations_bavaria(years)
    holidays_map = bavarian_holidays(years)
    ws.cell(2, 1, "Ferien")
    for col, day in enumerate(days, 2):
        if day.isoformat() in vacations:
            ws.cell(2, col, vacations[day.isoformat()]).fill = PatternFill(
                "solid", fgColor=VACATION_COLOR
            )
        elif day.isoformat() in holidays_map:
            ws.cell(2, col, holidays_map[day.isoformat()]).fill = PatternFill(
                "solid", fgColor=HOLIDAY_COLOR
            )
    row_no = 3
    for user in users:
        if user["id"] not in active_user_ids:
            continue
        ws.cell(row_no, 1, f"{user['first_name']} {user['last_name']}")
        for col, day in enumerate(days, 2):
            for code in entry_map.get((user["id"], day.isoformat()), []):
                ws.cell(row_no, col, code if current_user.has_role("admin") else "")
                ws.cell(row_no, col).fill = PatternFill(
                    "solid",
                    fgColor=fill_for(
                        current_user, user["id"], role_map[user["id"]], code
                    ),
                )
                break
        row_no += 1
    legend = row_no + 2
    ws.cell(legend, 1, "Legende")
    for idx, (code, (label, color)) in enumerate(ENTRY_TYPES.items(), legend + 1):
        ws.cell(idx, 1, code)
        ws.cell(idx, 2, label)
        ws.cell(idx, 3).fill = PatternFill(
            "solid",
            fgColor=SPECIAL_KR
            if code == "KR" and current_user.has_role("admin")
            else color,
        )
    xlsx_stream = BytesIO()
    wb.save(xlsx_stream)
    stream = macro_enabled_workbook(xlsx_stream)
    return send_file(
        stream,
        as_attachment=True,
        download_name=f"urlaubsuebersicht_{start_year}_{end_year}.xlsm",
        mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
    )


def macro_enabled_workbook(xlsx_stream: BytesIO) -> BytesIO:
    """Add the bundled vacation-layout VBA project to a generated workbook."""
    xlsx_stream.seek(0)
    result = BytesIO()
    with ZipFile(xlsx_stream, "r") as source, ZipFile(
        result, "w", ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                content = content.replace(
                    b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                    b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                )
                content = content.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/vbaProject.bin" '
                    b'ContentType="application/vnd.ms-office.vbaProject"/></Types>',
                )
            elif item.filename == "xl/_rels/workbook.xml.rels":
                content = content.replace(
                    b"</Relationships>",
                    b'<Relationship Id="rIdVbaProject" '
                    b'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
                    b'Target="vbaProject.bin"/></Relationships>',
                )
            target.writestr(item, content)
        target.writestr("xl/vbaProject.bin", bundled_vba_project())
    result.seek(0)
    return result


def bundled_vba_project() -> bytes:
    """Decode the text-only representation of the bundled VBA project."""
    return base64.b64decode(VBA_PROJECT_BASE64_PATH.read_text(encoding="ascii"))


@app.route("/upload", methods=["GET", "POST"])
@login_required
def upload():
    if not current_user.has_role("admin"):
        abort(403)
    if request.method == "POST":
        file = request.files["file"]
        filename = secure_filename(file.filename)
        Path(app.config["UPLOAD_FOLDER"]).mkdir(parents=True, exist_ok=True)
        path = Path(app.config["UPLOAD_FOLDER"]) / filename
        file.save(path)
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            import_excel(path)
            flash("Excel-Datei wurde importiert.")
            return redirect(url_for("index"))
        if suffix in {".doc", ".docx"}:
            imported = import_desksharing_word(path, date.today().year)
            flash(f"Word-Datei wurde importiert ({imported} Einträge).")
            return redirect(url_for("desksharing"))
        path.unlink(missing_ok=True)
        abort(400)
    return render_template("upload.html")


def extract_word_rows(path: Path) -> list[list[str]]:
    if path.suffix.lower() == ".docx" or path.read_bytes()[:2] == b"PK":
        document = Document(path)
        rows = []
        for block in document.iter_inner_content():
            if hasattr(block, "rows"):
                rows.extend(
                    [cell.text.strip() for cell in table_row.cells]
                    for table_row in block.rows
                )
            else:
                rows.extend([line.strip()] for line in block.text.splitlines())
        return [row for row in rows if any(row)]
    antiword = shutil.which("antiword")
    if antiword:
        result = subprocess.run(
            [antiword, str(path)], capture_output=True, check=False, text=True
        )
        if result.returncode == 0 and result.stdout.strip():
            return [
                line.split("\t") for line in result.stdout.splitlines() if line.strip()
            ]
    raw = path.read_bytes()
    decoded = raw.decode("utf-8", errors="ignore")
    if not decoded.strip() or decoded.count("\x00") > len(decoded) // 4:
        decoded = raw.decode("utf-16le", errors="ignore")
    lines = []
    for line in decoded.replace("\x00", "").splitlines():
        cleaned = "".join(character for character in line if character.isprintable())
        if cleaned.strip():
            lines.append(cleaned.split("\t"))
    return lines


def desksharing_user_map(conn) -> dict[str, int | None]:
    grouped = defaultdict(list)
    for user in get_desksharing_users(conn):
        grouped[user["first_name"].strip().casefold()].append(user["id"])
    return {
        first_name: user_ids[0] if len(user_ids) == 1 else None
        for first_name, user_ids in grouped.items()
    }


def desksharing_date(year: int, week: int, weekday: int) -> date | None:
    try:
        return date.fromisocalendar(year, week, weekday)
    except ValueError:
        return None


def import_desksharing_word(path: Path, year: int | None = None) -> int:
    rows = extract_word_rows(path)
    import_year = year or date.today().year
    current_week = None
    current_day = None
    current_status = None
    matrix_days: list[int | None] = []
    imported = 0
    with db() as conn:
        users = desksharing_user_map(conn)

        def save(first_name: str, weekday: int, status: str) -> None:
            nonlocal imported
            normalized_name = (
                re.sub(r"\s*\([^)]*\)\s*$", "", first_name).strip().casefold()
            )
            user_id = users.get(normalized_name)
            entry_date = (
                desksharing_date(import_year, current_week, weekday)
                if current_week
                else None
            )
            if not user_id or not entry_date:
                return
            conn.execute(
                """INSERT INTO desksharing_entries(
                    user_id, entry_date, status, created_by
                ) VALUES (?, ?, ?, ?)
                ON CONFLICT(user_id, entry_date) DO UPDATE SET
                    status = excluded.status, created_by = excluded.created_by""",
                (
                    user_id,
                    entry_date.isoformat(),
                    status,
                    getattr(current_user, "id", None),
                ),
            )
            imported += 1

        for row in rows:
            raw_cells = [cell.strip() if cell else "" for cell in row]
            cells = [cell for cell in raw_cells if cell]
            if not cells:
                continue
            joined = " ".join(cells)
            week_match = re.search(
                r"\b(?:KW|Kalenderwoche)\s*0?(\d{1,2})\b", joined, re.I
            )
            if week_match:
                current_week = int(week_match.group(1))
                current_day = None
                current_status = None
                if len(cells) == 1:
                    continue
            header_days = [
                DESKSHARING_WEEKDAYS.get(cell.strip(" .:").casefold())
                for cell in raw_cells[1:]
            ]
            if header_days and sum(day is not None for day in header_days) >= 2:
                matrix_days = header_days
                continue
            first_name_key = raw_cells[0].casefold()
            if matrix_days and users.get(first_name_key):
                for weekday, value in zip(matrix_days, raw_cells[1:], strict=False):
                    status = next(
                        (
                            name
                            for name in DESKSHARING_STATUSES
                            if name.casefold() == value.casefold()
                        ),
                        None,
                    )
                    if weekday and status:
                        save(raw_cells[0], weekday, status)
                continue
            for cell in cells:
                day = next(
                    (
                        weekday
                        for label, weekday in DESKSHARING_WEEKDAYS.items()
                        if re.search(rf"\b{re.escape(label)}\b", cell, re.I)
                    ),
                    None,
                )
                if day:
                    current_day = day
                status = next(
                    (
                        name
                        for name in DESKSHARING_STATUSES
                        if re.search(rf"\b{name}\b", cell, re.I)
                    ),
                    None,
                )
                if status:
                    current_status = status
                remainder = re.sub(
                    r"\b(?:KW|Kalenderwoche)\s*\d{1,2}\b", "", cell, flags=re.I
                )
                for label in DESKSHARING_WEEKDAYS:
                    remainder = re.sub(
                        rf"\b{re.escape(label)}\b", "", remainder, flags=re.I
                    )
                for name in DESKSHARING_STATUSES:
                    remainder = re.sub(rf"\b{name}\b", "", remainder, flags=re.I)
                if current_week and current_day and current_status:
                    for candidate in re.split(
                        r"[,;/]|\s+und\s+", remainder.strip(" :-"), flags=re.I
                    ):
                        if candidate.strip().casefold() in users:
                            save(candidate, current_day, current_status)
    return imported


def excel_fill_key(cell) -> tuple | None:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    color = fill.fgColor
    if color.type == "rgb" and color.rgb and color.rgb != "00000000":
        return ("rgb", color.rgb.upper())
    if color.type == "indexed" and color.indexed is not None:
        return ("indexed", color.indexed)
    if color.type == "theme" and color.theme is not None:
        return ("theme", color.theme, color.tint)
    return None


def build_legend_fill_map(ws) -> dict[tuple, str | None]:
    result = {}
    for row in ws.iter_rows():
        label = str(row[0].value).strip().casefold() if row[0].value else ""
        if label not in IMPORT_LEGEND_LABEL_TO_CODE:
            continue
        fill_key = excel_fill_key(row[0])
        if fill_key:
            result[fill_key] = IMPORT_LEGEND_LABEL_TO_CODE[label]
    return result


def parse_excel_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text[-10:], fmt).date()
            except ValueError:
                pass
    return None


GERMAN_MONTHS = {
    "jan": 1,
    "januar": 1,
    "februar": 2,
    "märz": 3,
    "maerz": 3,
    "april": 4,
    "mai": 5,
    "juni": 6,
    "juli": 7,
    "august": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "dezember": 12,
}


def reconstruct_dates_from_excel_headers(ws) -> tuple[int, dict[int, date]]:
    month_row = None
    day_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        label = str(row[0].value).strip().casefold() if row[0].value else ""
        if label == "monat":
            month_row = row[0].row
        elif label == "ferienzeit":
            day_row = row[0].row
    if not month_row or not day_row:
        return 1, {}
    dates = {}
    current_year = None
    current_month = None
    for col in range(2, ws.max_column + 1):
        marker = ws.cell(month_row, col).value
        if isinstance(marker, int) and 2000 <= marker <= 2100:
            current_year = marker
        elif isinstance(marker, str):
            month = GERMAN_MONTHS.get(marker.strip().casefold())
            if month:
                current_month = month
        parsed = parse_excel_date(ws.cell(day_row, col).value)
        if parsed:
            dates[col] = parsed
            current_year = parsed.year
            current_month = parsed.month
            continue
        day_value = ws.cell(day_row, col).value
        if current_year and current_month and isinstance(day_value, int):
            try:
                dates[col] = date(current_year, current_month, day_value)
            except ValueError:
                pass
    return day_row, dates


def find_excel_date_columns(ws) -> tuple[int, dict[int, date]]:
    best_row = 1
    best_dates: dict[int, date] = {}
    for row in ws.iter_rows():
        dates = {}
        for cell in row[1:]:
            parsed = parse_excel_date(cell.value)
            if parsed:
                dates[cell.column] = parsed
        if len(dates) > len(best_dates):
            best_row = row[0].row
            best_dates = dates
    reconstructed_row, reconstructed_dates = reconstruct_dates_from_excel_headers(ws)
    if len(reconstructed_dates) > len(best_dates):
        return reconstructed_row, reconstructed_dates
    return best_row, best_dates


def import_excel(path: Path) -> None:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    date_row, date_columns = find_excel_date_columns(ws)
    legend_fill_map = build_legend_fill_map(ws)
    actor_id = getattr(current_user, "id", None)
    with db() as conn:
        for row in ws.iter_rows(min_row=date_row + 1):
            row_label = str(row[0].value).strip() if row[0].value else ""
            if not row_label or row_label.lower() in IGNORED_IMPORT_NAMES:
                continue
            parts = row_label.split()
            first, last = (parts[0], " ".join(parts[1:]) or parts[0])
            username = f"{first}.{last}".lower().replace(" ", ".")
            user = conn.execute(
                "SELECT id FROM users WHERE username = ?", (username,)
            ).fetchone()
            user_id = (
                user["id"]
                if user
                else create_initial_user(conn, username, first, last, ["normal"])[0]
            )
            for cell in row[1:]:
                entry_date = date_columns.get(cell.column)
                raw_value = str(cell.value).strip() if cell.value else ""
                code = IMPORT_VALUE_TO_CODE_NORMALIZED.get(raw_value.casefold())
                if code is None and raw_value.replace("\xa0", "") == "":
                    code = legend_fill_map.get(excel_fill_key(cell))
                if code and entry_date:
                    conn.execute(
                        "INSERT OR IGNORE INTO entries(user_id, entry_date, code, created_by) VALUES (?, ?, ?, ?)",
                        (
                            user_id,
                            entry_date.isoformat(),
                            code,
                            actor_id,
                        ),
                    )


@app.context_processor
def inject_globals():
    return {"role_order": ROLE_ORDER, "entry_types": ENTRY_TYPES}


@app.before_request
def ensure_database():
    init_db()
    if (
        current_user.is_authenticated
        and session.get("pending_initial_user_id")
        and request.endpoint
        not in {"members", "acknowledge_initial_data", "logout", "static"}
    ):
        return redirect(url_for("members"))
    allowed = {"initial_login", "logout", "static"}
    if (
        current_user.is_authenticated
        and current_user.must_change_password
        and request.endpoint not in allowed
    ):
        return redirect(url_for("initial_login"))


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=4010, debug=True)
